import json
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from flask import Flask, request, jsonify, render_template
from flask import session as flask_session

app = Flask(__name__)
app.secret_key = 'e42b99233b821df3d7cc9ecdddf6439c'

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/search_user', methods=['GET'])
def search_user():
    user = request.args.get('user')
    session = create_session_from_cookies()
    if user:
        results = []
        if utilizador(session, results, user):
            return app.response_class(
                response=json.dumps(results, indent=4, ensure_ascii=False),
                mimetype='application/json'
            )
        else:
            return app.response_class(
                response=json.dumps({"error": "Failed to fetch user data"}, indent=4, ensure_ascii=False),
                mimetype='application/json'
            )
    return app.response_class(
        response=json.dumps({"error": "Sem parâmetro de utilizador"}, indent=4, ensure_ascii=False),
        mimetype='application/json'
    )

@app.route('/search_name', methods=['GET'])
def search_name():
    name = request.args.get('name')
    session = create_session_from_cookies()
    request.session = requests.Session()
    if name:
        results = []
        if nome(session, results, name):
            return app.response_class(
                response=json.dumps(results, indent=4, ensure_ascii=False),
                mimetype='application/json'
            )
        else:
            return app.response_class(
                response=json.dumps({"error": "Failed to fetch name data"}, indent=4, ensure_ascii=False),
                mimetype='application/json'
            )
    return app.response_class(
        response=json.dumps({"error": "Sem parâmetro de nome"}, indent=4, ensure_ascii=False),
        mimetype='application/json'
    )

@app.route('/search_printers', methods=['GET'])
def search_printers():
    location = request.args.get('location')
    if location:
        results = impressoras(location)
        return jsonify(results)
    return jsonify({"error": "Sem parâmetro de localização"}), 400

@app.route('/search_printer_by_ip', methods=['GET'])
def search_printer_by_ip():
    ip = request.args.get('ip')
    if ip:
        results = []
        wb = load_workbook('static/print.xlsx')
        sheet = wb.active
        row = 0
        found = False

        for i in range(53):
            row = i + 2
            ip_value = sheet["E" + str(row)].value
            if ip == ip_value:
                found = True
                break

        local = str(sheet['C' + str(row)].value) if "N/A" in str(sheet['D' + str(row)].value) else str(sheet['D' + str(row)].value)

        if found:
            results.append({
                "Hostname": str(sheet['A' + str(row)].value),
                "Local": local,
                "IP": str(sheet['E' + str(row)].value),
                "User": str(sheet['F' + str(row)].value),
                "Password": str(sheet['G' + str(row)].value)
            })
        else:
            results.append({"error": "Printer with this IP not found"})
        return jsonify(results)
    return jsonify({"error": "Sem parâmetro de endereço IP"}), 400

@app.route('/printers_by_location', methods=['GET'])
def impressoras():
    selected_location = request.args.get('location')
    if selected_location:
        output = []
        locais = {
            "2º PISO": [],
            "3º PISO": [],
            "4º PISO": [],
            "6º PISO": [],
            "IRAE": [],
            "RVAH": [],
            "NOF": [],
            "RVH": [],
            "IRT - Faial": [],
            "IRT - Terceira": [],
            "NOT": []
        }
        wb = load_workbook('static/print.xlsx')
        sheet = wb.active
        for i in range(53):
            location = sheet["D" + str(i+2)].value
            data = {
                "Hostname": sheet["A" + str(i+2)].value,
                "Local": sheet["B" + str(i+2)].value,
                "IP": sheet["E" + str(i+2)].value,
                "User": sheet["F" + str(i+2)].value,
                "Password": sheet["G" + str(i+2)].value
            }
            if "N/A" in str(sheet["D" + str(i+2)].value):
                data["Serviço"] = sheet["C" + str(i+2)].value
            if location == "2º PISO":
                locais["2º PISO"].append(data)
            elif location == "3º PISO":
                locais["3º PISO"].append(data)
            elif location == "4º PISO":
                locais["4º PISO"].append(data)
            elif location == "6º PISO":
                locais["6º PISO"].append(data)
            elif location == "N/A" and "IRAE" in str(sheet["C" + str(i+2)].value):
                locais["IRAE"].append(data)
            elif location == "N/A" and "RVAH" in str(sheet["C" + str(i+2)].value):
                locais["RVAH"].append(data)
            elif location == "N/A" and "NOF" in str(sheet["C" + str(i+2)].value):
                locais["NOF"].append(data)
            elif location == "N/A" and "RVH" in str(sheet["C" + str(i+2)].value):
                locais["RVH"].append(data)
            elif location == "N/A" and "IRT - Faial" in str(sheet["C" + str(i+2)].value):
                locais["IRT - Faial"].append(data)
            elif location == "N/A" and "IRT - Terceira" in str(sheet["C" + str(i+2)].value):
                locais["IRT - Terceira"].append(data)
            elif location == "N/A" and "NOT" in str(sheet["C" + str(i+2)].value):
                locais["NOT"].append(data)
        # Return only printers from the selected location
        if selected_location in locais:
            output = locais[selected_location]
            return jsonify(output)
        else:
            return jsonify({"error": "Invalid location"}), 400
    return jsonify({"error": "Sem parâmetro de localização"}), 400

@app.route('/login', methods=['POST'])
def login():
    session = requests.Session()
    data = request.get_json()  # Parse JSON data from the request
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400
    
    login_url = "https://myteam.azores.gov.pt/admin/site/login"

    response = session.get(login_url)
    if response.status_code != 200:
        return jsonify({"error": "Failed to access login page"}), 500

    soup = BeautifulSoup(response.text, "html.parser")
    csrf_token = soup.find("input", {"name": "_csrf-backend"})["value"]

    login_data = {
        "_csrf-backend": csrf_token.strip("=="),
        "LoginForm[username]": username,
        "LoginForm[password]": password,
        "LoginForm[rememberMe]": "0",
        "login-button": ""
    }

    print("Login data:", login_data)

    response = session.post(login_url, data=login_data, allow_redirects=True)

    # Debugging: Log the response content
    print("Login Response Content:", response.url)

    # Check if login was successful
    if response.url == "https://myteam.azores.gov.pt/admin/":
        flask_session['sessions'] = session.cookies.get_dict()
        return jsonify({"message": "Login successful", "success": True}), 200
    elif response.url == "https://myteam.azores.gov.pt/admin/site/login":
        return jsonify({"message": "Invalid username or password.", "success": False}), 401
    else:
        return jsonify({"message": "Unable to process login.", "success": False}), 500

def create_session_from_cookies():
    session = requests.Session()
    if 'sessions' in flask_session:
        session.cookies.update(flask_session['sessions'])
    return session 

def utilizador(session, results, user):
    target_url = f"https://myteam.azores.gov.pt/admin/users/index?UsersSearch%5Bcolaborador%5D=&UsersSearch%5Busername%5D={user}&UsersSearch%5Bvoip%5D=&UsersSearch%5Bid_ilha%5D=&UsersSearch%5Bid_edificio%5D=&UsersSearch%5Bstatus%5D=&UsersSearch%5Badmin_access%5D=&UsersSearch%5Bid_organismo%5D=&UsersSearch%5Bid_servico%5D="
    response = session.get(target_url)
    print("Response URL:", response.url)  # Debugging: Log the response URL
    if response.url != target_url:
        return False
    soup = BeautifulSoup(response.text, 'html.parser')
    p0 = soup.find('div', id='p0')
    w1 = p0.find('div', id='w1')
    if w1:
        table = w1.find('table', class_='table table-striped table-bordered')
        if table:
            tbody = table.find('tbody')
            if tbody:
                tr = tbody.find_all('tr')
                for i in tr:
                    td = i.find_all('td')
                    print(td[10].find('span').text.strip())  # Debugging: Log the status
                    if len(td) > 2:
                        if td[10].find('span').text.strip() != "Desativado":
                            results.append({
                                "Colaborador": td[1].text.strip(),
                                "Username": td[2].text.strip(),
                                "VoIP": td[3].text.strip(),
                                "Email": td[4].find('a')['href'].split('mailto:')[1].strip(),
                                "Serviço": td[6].text.strip(),
                                "Gabinete": td[8].text.strip() if td[8].text.strip() != "(não definido)" else "Não definido"
                        })
                        else:
                            results.append({"error": "O colaborador foi desativado"})
                # If no valid rows were found, append an error message
                if not results:
                    results.append({"error": "O utilizador não existe!"})
            else:
                results.append({"error": "O utilizador não existe!"})
        else:
            results.append({"error": "O utilizador não existe!"})
    else:
        results.append({"error": "O utilizador não existe!"})
    return True

def nome(session, results, user):
    target_url = f"https://myteam.azores.gov.pt/admin/users/index?UsersSearch%5Bcolaborador%5D={user.replace(' ', '+')}"
    response = session.get(target_url)
    if response.status_code != 200:
        return False
    soup = BeautifulSoup(response.text, 'html.parser')
    p0 = soup.find('div', id='p0')
    w1 = p0.find('div', id='w1') if p0 else None
    if w1:
        table = w1.find('table', class_='table table-striped table-bordered')
        if table:
            tbody = table.find('tbody')
            if tbody:
                tr = tbody.find_all('tr')
                for i in tr:
                    td = i.find_all('td')
                    if len(td) > 2:
                        if td[10].find('span').text.strip() != "Desativado":
                            results.append({
                                "Colaborador": td[1].text.strip(),
                                "Username": td[2].text.strip(),
                                "VoIP": td[3].text.strip(),
                                "Email": td[4].find('a')['href'].split('mailto:')[1].strip(),
                                "Serviço": td[6].text.strip(),
                                "Gabinete": td[8].text.strip() if td[8].text.strip() != "(não definido)" else "Não definido"
                            })
                # If no valid rows were found, append an error message
                if not results:
                    results.append({"error": "O utilizador não existe!"})
            else:
                results.append({"error": "O utilizador não existe!"})
        else:
            results.append({"error": "O utilizador não existe!"})
    else:
        results.append({"error": "O utilizador não existe!"})
    return True

app.run("0.0.0.0", 5000, debug=True)
